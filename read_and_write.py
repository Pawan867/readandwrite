from openpyxl import load_workbook

wb = load_workbook('task.xlsx')
sheet1 = wb.get_sheet_by_name('Pawan')
sheet2 = wb.get_sheet_by_name('Acharya')

for i in range(1, 11):
    for j in range(1, sheet1.max_column+1):
        sheet2.cell(row=i, column=j).value = sheet1.cell(row=i, column=j).value

wb.save('task.xlsx')
